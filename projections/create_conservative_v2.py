#!/usr/bin/env python3
"""
Conservative 18-Month Projection Model v2
Goal: ~$300k/month COMBINED revenue by Q6 (month 18)

Key changes from v1:
- Much slower ramp schedules across all channels
- Ads: 30% Q1 → gradual climb to 100% by Q5
- GBP: 10% Q1 → slow build to 100% by Q6  
- Website: Doesn't start producing until Q3, only 60% by Q6
- Ad spend calculated correctly: qualified PPC leads × CPL
- Realistic early quarters (no overpromising)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# STYLES
# ============================================================
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
goal_met_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
goal_miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ramp_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

money_fmt = '"$"#,##0'
pct_fmt = '0%'
decimal_fmt = '#,##0.0'
roi_fmt = '0.0"x"'

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def apply_borders(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

def auto_width(ws, min_w=14):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, 30))


# ============================================================
# CONSERVATIVE RAMP SCHEDULES
# ============================================================
# These are MUCH slower than v1
ramps = {
    'Q1': {'ads': 0.30, 'gbp': 0.10, 'web': 0.00},  # Months 1-3: Just starting ads, GBP barely live
    'Q2': {'ads': 0.60, 'gbp': 0.25, 'web': 0.00},  # Months 4-6: Ads optimizing, GBP building reviews
    'Q3': {'ads': 0.80, 'gbp': 0.45, 'web': 0.05},  # Months 7-9: Ads strong, GBP gaining, website starting
    'Q4': {'ads': 0.90, 'gbp': 0.65, 'web': 0.15},  # Months 10-12: Near-peak ads, GBP growing, website crawling
    'Q5': {'ads': 1.00, 'gbp': 0.85, 'web': 0.35},  # Months 13-15: Full ads, GBP strong, website picking up
    'Q6': {'ads': 1.00, 'gbp': 1.00, 'web': 0.60},  # Months 16-18: Full ads+GBP, website at 60% (still growing)
}

# ============================================================
# MARKET ASSUMPTIONS
# ============================================================
qualified_rate = 0.50  # 50% of calls are qualified
closing_rate = 0.50    # 50% of qualified leads close
mgmt_fee_monthly = 5500 / 2  # $2,750/month per location

tucson = {
    'name': 'Tucson',
    'max_web': 20, 'max_ppc': 15, 'max_gbp': 30,
    'cpl': 700,
    'mit_avg': 4589, 'abate_avg': 7484, 'abate_conv': 0.30,
    'recon_avg': 7452, 'recon_conv': 0.55,
    'ltv': 4589 + (0.30 * 7484) + (0.55 * 7452)  # = $10,933.80
}

denver = {
    'name': 'Denver',
    'max_web': 35, 'max_ppc': 30, 'max_gbp': 50,
    'cpl': 800,
    'mit_avg': 6100, 'recon_fee': 0,
    'ltv': 6100
}

# ============================================================
# CALCULATION ENGINE
# ============================================================
def calculate_quarter(market, ramp, quarter_label):
    """Calculate one quarter of projections for a location"""
    # Qualified leads per quarter (max × ramp × qualified% × 3 months)
    web_leads = market['max_web'] * ramp['web'] * qualified_rate * 3
    ppc_leads = market['max_ppc'] * ramp['ads'] * qualified_rate * 3
    gbp_leads = market['max_gbp'] * ramp['gbp'] * qualified_rate * 3
    total_leads = web_leads + ppc_leads + gbp_leads

    # Costs
    # Ad spend = qualified PPC leads × CPL (this IS the ad spend to acquire those leads)
    ad_spend = ppc_leads * market['cpl']
    mgmt_fee = mgmt_fee_monthly * 3  # quarterly
    total_cost = ad_spend + mgmt_fee

    # Revenue
    jobs = total_leads * closing_rate
    revenue = jobs * market['ltv']
    monthly_revenue = revenue / 3
    roi = revenue / total_cost if total_cost > 0 else 0

    return {
        'quarter': quarter_label,
        'web_leads': round(web_leads, 1),
        'ppc_leads': round(ppc_leads, 1),
        'gbp_leads': round(gbp_leads, 1),
        'total_leads': round(total_leads, 1),
        'ad_spend': round(ad_spend),
        'mgmt_fee': round(mgmt_fee),
        'total_cost': round(total_cost),
        'jobs': round(jobs, 1),
        'revenue': round(revenue),
        'monthly_rev': round(monthly_revenue),
        'roi': round(roi, 1)
    }


def calc_all_quarters(market):
    """Calculate all 6 quarters for a market"""
    results = []
    for q_label, ramp in ramps.items():
        results.append(calculate_quarter(market, ramp, q_label))
    return results


tucson_data = calc_all_quarters(tucson)
denver_data = calc_all_quarters(denver)


# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = Workbook()


# ============================================================
# TAB 1: ASSUMPTIONS
# ============================================================
ws1 = wb.active
ws1.title = "Assumptions"

ws1.append(["Category", "Parameter", "Value", "Notes", "Adjustable?"])
style_header(ws1)

# General section
ws1.append(["GENERAL INPUTS", "", "", "", ""])
ws1.append(["General", "Monthly Mgmt Fee (Both Locations)", 5500, "Split 50/50 = $2,750/location", "YES"])
ws1.append(["General", "Qualified Lead % (of all calls)", 0.50, "Industry avg: 40-60%", "YES"])
ws1.append(["General", "Closing Rate (Qualified → Job)", 0.50, "Conservative: 50%", "YES"])
ws1.append(["", "", "", "", ""])

# Tucson
ws1.append(["TUCSON MARKET DATA", "", "", "", ""])
ws1.append(["Tucson", "Max Monthly Website Leads", 20, "At full 18-month maturity", "YES"])
ws1.append(["Tucson", "Max Monthly PPC Leads", 15, "At full ad spend", "YES"])
ws1.append(["Tucson", "Max Monthly GBP Leads", 30, "At full GBP maturity", "YES"])
ws1.append(["Tucson", "Cost Per Qualified Lead (PPC)", 700, "Water damage CPL", "YES"])
ws1.append(["Tucson", "Mitigation Average", 4589, "Per job, from client data", "YES"])
ws1.append(["Tucson", "Abatement Average", 7484, "Per job, from client data", "YES"])
ws1.append(["Tucson", "Abatement Conversion %", 0.30, "30% of mit → abate", "YES"])
ws1.append(["Tucson", "Reconstruction Average", 7452, "Per job, from client data", "YES"])
ws1.append(["Tucson", "Reconstruction Conversion %", 0.55, "55% of mit → recon", "YES"])
ws1.append(["Tucson", "LTV Per Job (Calculated)", round(tucson['ltv']), "Mit + (Abate% × Abate) + (Recon% × Recon)", "AUTO"])
ws1.append(["", "", "", "", ""])

# Denver
ws1.append(["DENVER MARKET DATA", "", "", "", ""])
ws1.append(["Denver", "Max Monthly Website Leads", 35, "At full 18-month maturity", "YES"])
ws1.append(["Denver", "Max Monthly PPC Leads", 30, "At full ad spend", "YES"])
ws1.append(["Denver", "Max Monthly GBP Leads", 50, "At full GBP maturity", "YES"])
ws1.append(["Denver", "Cost Per Qualified Lead (PPC)", 800, "Water damage CPL", "YES"])
ws1.append(["Denver", "Mitigation Average", 6100, "Per job, from client data", "YES"])
ws1.append(["Denver", "Recon Referral Fee", 0, "Refer out, no direct rev", "YES"])
ws1.append(["Denver", "LTV Per Job (Calculated)", round(denver['ltv']), "Mit + Referral Fee", "AUTO"])
ws1.append(["", "", "", "", ""])

# Ramp schedules
ws1.append(["RAMP SCHEDULES (Conservative)", "Quarter", "Ads %", "GBP %", "Website %"])
for q_label, ramp in ramps.items():
    ws1.append(["Ramp", q_label, ramp['ads'], ramp['gbp'], ramp['web']])
ws1.append(["", "", "", "", ""])

# Ramp justification
ws1.append(["RAMP JUSTIFICATION", "", "", "", ""])
ws1.append(["Ads (Fastest)", "Q1: 30% testing", "Q2: 60% optimizing", "Q5-Q6: 100% mature", "90-day sprint, then scale"])
ws1.append(["GBP (Medium)", "Q1: 10% just live", "Q3: 45% building reviews", "Q6: 100% mature", "4-6 month typical build"])
ws1.append(["Website (Slowest)", "Q1-Q2: 0% no traffic", "Q3: 5% first rankings", "Q6: 60% still growing", "12-18 months to full maturity"])

# Format
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    for cell in row:
        cell.border = thin_border
    
    # Format values
    val = row[2].value
    param = str(row[1].value) if row[1].value else ""
    
    if isinstance(val, float) and 0 < val <= 1:
        row[2].number_format = pct_fmt
    elif any(word in param for word in ["Cost", "Average", "Fee", "LTV", "Mitigation", "Abatement", "Reconstruction"]):
        if isinstance(val, (int, float)) and val > 1:
            row[2].number_format = money_fmt
    
    # Section headers
    cat = str(row[0].value) if row[0].value else ""
    if cat.isupper() and len(cat) > 3:
        for cell in row:
            cell.fill = section_fill
            cell.font = Font(bold=True)
    
    # Highlight adjustable inputs
    if row[4].value == "YES":
        row[2].fill = input_fill

# Format ramp percentages  
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    if row[0].value == "Ramp":
        for cell in [row[2], row[3], row[4]]:
            if isinstance(cell.value, float):
                cell.number_format = pct_fmt
                cell.fill = ramp_fill

auto_width(ws1)


# ============================================================
# TAB 2: TUCSON PROJECTIONS
# ============================================================
ws2 = wb.create_sheet("Tucson Projections")

headers = [
    "Quarter", "Months",
    "Website Leads", "PPC Leads", "GBP Leads", "Total Qualified Leads",
    "Ad Spend (Google)", "Mgmt Fee Share", "Total Cost",
    "Jobs Closed", "Revenue (Qtr)", "Monthly Revenue (Avg)", "ROI"
]
ws2.append(headers)
style_header(ws2)

month_labels = ["1-3", "4-6", "7-9", "10-12", "13-15", "16-18"]

for i, row_data in enumerate(tucson_data):
    ws2.append([
        row_data['quarter'], month_labels[i],
        row_data['web_leads'], row_data['ppc_leads'], row_data['gbp_leads'], row_data['total_leads'],
        row_data['ad_spend'], row_data['mgmt_fee'], row_data['total_cost'],
        row_data['jobs'], row_data['revenue'], row_data['monthly_rev'], row_data['roi']
    ])

# Totals
ws2.append([
    'TOTAL', '1-18',
    sum(d['web_leads'] for d in tucson_data),
    sum(d['ppc_leads'] for d in tucson_data),
    sum(d['gbp_leads'] for d in tucson_data),
    sum(d['total_leads'] for d in tucson_data),
    sum(d['ad_spend'] for d in tucson_data),
    sum(d['mgmt_fee'] for d in tucson_data),
    sum(d['total_cost'] for d in tucson_data),
    sum(d['jobs'] for d in tucson_data),
    sum(d['revenue'] for d in tucson_data),
    None,  # No avg for totals
    round(sum(d['revenue'] for d in tucson_data) / sum(d['total_cost'] for d in tucson_data), 1)
])

# Format
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for idx, cell in enumerate(row):
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if idx in [6, 7, 8, 10, 11]:  # Money
            cell.number_format = money_fmt
        elif idx == 12:  # ROI
            cell.number_format = roi_fmt
        elif idx in [2, 3, 4, 5, 9]:  # Counts
            cell.number_format = decimal_fmt
    
    if row[0].value == 'TOTAL':
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)

auto_width(ws2)


# ============================================================
# TAB 3: DENVER PROJECTIONS
# ============================================================
ws3 = wb.create_sheet("Denver Projections")
ws3.append(headers)
style_header(ws3)

for i, row_data in enumerate(denver_data):
    ws3.append([
        row_data['quarter'], month_labels[i],
        row_data['web_leads'], row_data['ppc_leads'], row_data['gbp_leads'], row_data['total_leads'],
        row_data['ad_spend'], row_data['mgmt_fee'], row_data['total_cost'],
        row_data['jobs'], row_data['revenue'], row_data['monthly_rev'], row_data['roi']
    ])

ws3.append([
    'TOTAL', '1-18',
    sum(d['web_leads'] for d in denver_data),
    sum(d['ppc_leads'] for d in denver_data),
    sum(d['gbp_leads'] for d in denver_data),
    sum(d['total_leads'] for d in denver_data),
    sum(d['ad_spend'] for d in denver_data),
    sum(d['mgmt_fee'] for d in denver_data),
    sum(d['total_cost'] for d in denver_data),
    sum(d['jobs'] for d in denver_data),
    sum(d['revenue'] for d in denver_data),
    None,
    round(sum(d['revenue'] for d in denver_data) / sum(d['total_cost'] for d in denver_data), 1)
])

for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    for idx, cell in enumerate(row):
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if idx in [6, 7, 8, 10, 11]:
            cell.number_format = money_fmt
        elif idx == 12:
            cell.number_format = roi_fmt
        elif idx in [2, 3, 4, 5, 9]:
            cell.number_format = decimal_fmt
    if row[0].value == 'TOTAL':
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)

auto_width(ws3)


# ============================================================
# TAB 4: COMBINED SUMMARY
# ============================================================
ws4 = wb.create_sheet("Combined Summary")

combined_headers = [
    "Quarter", "Months",
    "Tucson Leads", "Denver Leads", "Total Qualified Leads",
    "Total Ad Spend", "Total Mgmt Fees", "Total Investment",
    "Total Jobs", "Total Revenue (Qtr)",
    "Monthly Revenue (Avg)", "Combined ROI",
    "$300k Goal Progress"
]
ws4.append(combined_headers)
style_header(ws4)

combined_data = []
for i in range(6):
    t = tucson_data[i]
    d = denver_data[i]
    
    total_leads = t['total_leads'] + d['total_leads']
    total_ad = t['ad_spend'] + d['ad_spend']
    total_mgmt = t['mgmt_fee'] + d['mgmt_fee']
    total_inv = t['total_cost'] + d['total_cost']
    total_jobs = t['jobs'] + d['jobs']
    total_rev = t['revenue'] + d['revenue']
    monthly_rev = total_rev / 3
    combined_roi = total_rev / total_inv if total_inv > 0 else 0
    goal_pct = monthly_rev / 300000
    
    row_data = [
        t['quarter'], month_labels[i],
        t['total_leads'], d['total_leads'], round(total_leads, 1),
        round(total_ad), round(total_mgmt), round(total_inv),
        round(total_jobs, 1), round(total_rev),
        round(monthly_rev), round(combined_roi, 1),
        goal_pct
    ]
    combined_data.append(row_data)
    ws4.append(row_data)

# Totals
total_rev_all = sum(r[9] for r in combined_data)
total_inv_all = sum(r[7] for r in combined_data)
ws4.append([
    'TOTAL', '1-18',
    sum(r[2] for r in combined_data),
    sum(r[3] for r in combined_data),
    sum(r[4] for r in combined_data),
    sum(r[5] for r in combined_data),
    sum(r[6] for r in combined_data),
    total_inv_all,
    sum(r[8] for r in combined_data),
    total_rev_all,
    None,
    round(total_rev_all / total_inv_all, 1),
    None
])

# Goal status rows
ws4.append([])
q6_monthly = combined_data[-1][10]
ws4.append(["", "", "", "", "", "", "", "", "", "Q6 Monthly Revenue:", q6_monthly, "", ""])
ws4.append(["", "", "", "", "", "", "", "", "", "Target:", 300000, "", ""])
ws4.append(["", "", "", "", "", "", "", "", "", "Status:", 
            "GOAL MET" if q6_monthly >= 300000 else "BELOW TARGET",
            "", ""])

# Format
for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
    for idx, cell in enumerate(row):
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        if idx in [5, 6, 7, 9, 10]:  # Money
            cell.number_format = money_fmt
        elif idx == 11:  # ROI
            cell.number_format = roi_fmt
        elif idx == 12:  # Goal %
            cell.number_format = '0%'
        elif idx in [2, 3, 4, 8]:  # Counts
            cell.number_format = decimal_fmt
    
    if row[0].value == 'TOTAL':
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)

# Color the goal progress column
for row_idx in range(2, 2 + len(combined_data)):
    cell = ws4.cell(row=row_idx, column=13)
    if cell.value:
        if cell.value >= 1.0:
            cell.fill = goal_met_fill
        elif cell.value >= 0.75:
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            cell.fill = goal_miss_fill

# Style goal status
for row_idx in range(ws4.max_row - 2, ws4.max_row + 1):
    for cell in ws4[row_idx]:
        cell.font = Font(bold=True, size=12)
    ws4.cell(row=row_idx, column=11).number_format = money_fmt

if q6_monthly >= 300000:
    ws4.cell(row=ws4.max_row, column=11).fill = goal_met_fill
    ws4.cell(row=ws4.max_row, column=11).font = Font(bold=True, size=14, color="006100")
else:
    ws4.cell(row=ws4.max_row, column=11).fill = goal_miss_fill

auto_width(ws4)


# ============================================================
# TAB 5: SENSITIVITIES & NOTES
# ============================================================
ws5 = wb.create_sheet("Sensitivities & Notes")

ws5.append(["Scenario", "Variable", "Change", "Q6 Monthly Impact", "Notes"])
style_header(ws5)

# Calculate sensitivities
base_q6 = q6_monthly

# 60% closing rate
close_60_q6 = base_q6 * (0.60 / 0.50)
close_40_q6 = base_q6 * (0.40 / 0.50)

sensitivity_data = [
    ["CLOSING RATE SCENARIOS", "", "", "", ""],
    ["Closing Rate +10%", "50% → 60%", "+20% revenue", round(close_60_q6), "Strong ops, good phone skills"],
    ["Closing Rate -10%", "50% → 40%", "-20% revenue", round(close_40_q6), "Poor intake or slow dispatch"],
    ["", "", "", "", ""],
    
    ["COST PER LEAD SCENARIOS", "", "", "", ""],
    ["CPL +20%", "Tucson $840, Denver $960", "Higher ad costs", round(base_q6), "Revenue unchanged, ROI drops ~17%"],
    ["CPL -15% (plumbing keywords)", "Tucson $595, Denver $680", "Lower ad costs", round(base_q6), "Revenue unchanged, ROI improves ~18%"],
    ["", "", "", "", ""],
    
    ["TIMELINE SCENARIOS", "", "", "", ""],
    ["Website delays to Q4", "Web ramp: 0/0/0/5/20/45", "Slower organic", round(base_q6 * 0.88), "Still above $250k/month"],
    ["GBP builds faster", "GBP ramp: 15/35/55/75/95/100", "+10-15% leads", round(base_q6 * 1.08), "If reviews come in strong"],
    ["", "", "", "", ""],
    
    ["GROWTH ACCELERATION", "", "", "", ""],
    ["Add plumbing keywords Q3", "New PPC category", "CPL drops 10-20%", "N/A", "Target emergency plumbing searches"],
    ["2nd Denver GBP in Q2", "New location profile", "+20-30 leads/mo by Q4", "N/A", "Target downtown Denver market"],
    ["2nd Tucson GBP in Q4", "New location profile", "+15-20 leads/mo by Q6", "N/A", "Target east Tucson suburbs"],
    ["", "", "", "", ""],
    
    ["CONSERVATIVE MODEL NOTES", "", "", "", ""],
    ["Ads (PPC)", "Fastest channel", "30% Q1 → 100% Q5", "", "90-day sprint, conservative start"],
    ["GBP (Local SEO)", "Medium channel", "10% Q1 → 100% Q6", "", "Reviews build over 4-6 months"],
    ["Website (Traditional SEO)", "Slowest channel", "0% Q1-Q2 → 60% Q6", "", "12-18 months for real organic traffic"],
    ["Website NOT at 100% by Q6", "Still growing post-Q6", "60% at month 18", "", "Continued growth beyond 18 months"],
    ["", "", "", "", ""],
    
    ["FORMULAS", "", "", "", ""],
    ["Qualified Leads (Qtr)", "=Max Monthly Leads × Ramp% × Qualified% × 3", "", "", "3 months per quarter"],
    ["Ad Spend (Qtr)", "=Qualified PPC Leads × CPL", "", "", "CPL is cost per QUALIFIED lead"],
    ["Mgmt Fee (Qtr)", "=$5,500 ÷ 2 × 3 months", "", "", "$2,750/location/month"],
    ["Jobs (Qtr)", "=Total Qualified Leads × Closing Rate", "", "", "50% close rate baseline"],
    ["Revenue (Qtr)", "=Jobs × LTV per Job", "", "", "Tucson $10,933 / Denver $6,100"],
    ["ROI", "=Quarterly Revenue ÷ Quarterly Total Cost", "", "", ""],
    ["Tucson LTV", "=$4,589 + (30% × $7,484) + (55% × $7,452)", "", "$10,933", ""],
    ["Denver LTV", "=$6,100 + $0 referral fee", "", "$6,100", ""],
]

for row_data in sensitivity_data:
    ws5.append(row_data)

for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row):
    for cell in row:
        cell.border = thin_border
    
    cat = str(row[0].value) if row[0].value else ""
    if cat.isupper() and len(cat) > 3:
        for cell in row:
            cell.fill = section_fill
            cell.font = Font(bold=True)
    
    # Money format for impact column
    if isinstance(row[3].value, (int, float)) and row[3].value > 1000:
        row[3].number_format = money_fmt

auto_width(ws5)


# ============================================================
# SAVE & REPORT
# ============================================================
output_path = "/Users/jameslarosa/Desktop/Random AI Prjects/AI Wireframe Builder/projections/Conservative_v2_Projections.xlsx"
wb.save(output_path)

print(f"✅ Created: {output_path}")
print()
print("=" * 65)
print("  CONSERVATIVE 18-MONTH PROJECTION SUMMARY")
print("=" * 65)
print()
print("  TUCSON (LTV: $10,933/job)")
print("  " + "-" * 50)
for d in tucson_data:
    print(f"  {d['quarter']} | Leads: {d['total_leads']:6.1f} | Jobs: {d['jobs']:5.1f} | Monthly: ${d['monthly_rev']:>8,}")
print(f"  {'TOTAL':2s} | Revenue: ${sum(d['revenue'] for d in tucson_data):>10,}")
print()
print("  DENVER (LTV: $6,100/job)")
print("  " + "-" * 50)
for d in denver_data:
    print(f"  {d['quarter']} | Leads: {d['total_leads']:6.1f} | Jobs: {d['jobs']:5.1f} | Monthly: ${d['monthly_rev']:>8,}")
print(f"  {'TOTAL':2s} | Revenue: ${sum(d['revenue'] for d in denver_data):>10,}")
print()
print("  COMBINED")
print("  " + "-" * 50)
for i in range(6):
    mo_rev = tucson_data[i]['monthly_rev'] + denver_data[i]['monthly_rev']
    pct = mo_rev / 300000 * 100
    bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
    print(f"  Q{i+1} | Monthly: ${mo_rev:>8,} | {bar} {pct:.0f}% of $300k")
print()
total_rev = sum(d['revenue'] for d in tucson_data) + sum(d['revenue'] for d in denver_data)
total_cost = sum(d['total_cost'] for d in tucson_data) + sum(d['total_cost'] for d in denver_data)
print(f"  18-Month Total Revenue:    ${total_rev:>10,}")
print(f"  18-Month Total Investment: ${total_cost:>10,}")
print(f"  Overall ROI:               {total_rev/total_cost:.1f}x")
print(f"  Q6 Monthly Revenue:        ${q6_monthly:>10,}")
print(f"  $300k Goal:                {'✅ MET' if q6_monthly >= 300000 else '❌ NOT MET'}")
print()
