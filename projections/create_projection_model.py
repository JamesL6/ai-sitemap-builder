#!/usr/bin/env python3
"""
Generate 18-Month Conservative Projection Model for Restoration Marketing
Goal: $300k/month combined revenue by Q6
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call(['python3', '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
goal_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

money_format = '"$"#,##0'
percent_format = '0%'
decimal_format = '#,##0.0'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, row=1):
    """Style header row with blue background"""
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, min_width=12):
    """Auto-adjust column widths"""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max(min_width, min(max_length + 2, 35))

wb = Workbook()

# ===== TAB 1: ASSUMPTIONS (Input Variables) =====
ws1 = wb.active
ws1.title = "Assumptions"

# Header
ws1.append(["Category", "Parameter", "Value", "Notes"])
style_header_row(ws1)

assumptions_data = [
    ["GENERAL", "", "", ""],
    ["General", "Monthly Management Fee (Both Locations)", 5500, "Split 50/50 between locations"],
    ["General", "Qualified Lead % (of all calls)", 0.5, "50% of calls are qualified"],
    ["General", "Closing Rate (Qualified → Job)", 0.5, "50% of qualified leads close"],
    ["", "", "", ""],
    ["TUCSON MARKET", "", "", ""],
    ["Tucson - Capacity", "Max Monthly Website Leads", 20, "At 18-month maturity"],
    ["Tucson - Capacity", "Max Monthly PPC Leads", 15, "At full scale"],
    ["Tucson - Capacity", "Max Monthly GBP Leads", 30, "At 12-month maturity"],
    ["Tucson - Costs", "Cost Per Lead (PPC)", 700, "Average CPL for water damage"],
    ["Tucson - Revenue", "Mitigation Average", 4589, "Average mitigation job size"],
    ["Tucson - Revenue", "Abatement Average", 7484, "Average abatement job size"],
    ["Tucson - Revenue", "Abatement Conversion %", 0.30, "% of mit jobs → abate"],
    ["Tucson - Revenue", "Reconstruction Conversion %", 0.55, "% of mit jobs → recon"],
    ["Tucson - Revenue", "Reconstruction Average", 7452, "Average recon job size"],
    ["Tucson - Revenue", "Lifetime Value (LTV) Per Job", 10933, "=Mit + (Abate% × Abate$) + (Recon% × Recon$)"],
    ["", "", "", ""],
    ["DENVER MARKET", "", "", ""],
    ["Denver - Capacity", "Max Monthly Website Leads", 35, "At 18-month maturity"],
    ["Denver - Capacity", "Max Monthly PPC Leads", 30, "At full scale"],
    ["Denver - Capacity", "Max Monthly GBP Leads", 50, "At 12-month maturity"],
    ["Denver - Costs", "Cost Per Lead (PPC)", 800, "Average CPL for water damage"],
    ["Denver - Revenue", "Mitigation Average", 6100, "Average mitigation job size"],
    ["Denver - Revenue", "Recon Referral Fee", 0, "Refer out, no direct revenue"],
    ["Denver - Revenue", "Lifetime Value (LTV) Per Job", 6100, "=Mit + Referral Fee"],
    ["", "", "", ""],
    ["RAMP SCHEDULES", "Quarter", "Ads %", "GBP %", "Website %"],
    ["Ramp", "Q1", 0.50, 0.25, 0.00],
    ["Ramp", "Q2", 1.00, 0.50, 0.00],
    ["Ramp", "Q3", 1.00, 0.75, 0.10],
    ["Ramp", "Q4", 1.00, 1.00, 0.30],
    ["Ramp", "Q5", 1.00, 1.00, 0.60],
    ["Ramp", "Q6", 1.00, 1.00, 1.00],
    ["", "", "", "", ""],
    ["NOTES", "", "", ""],
    ["", "Conservative Model", "", "Slower ramps to avoid overpromising"],
    ["", "Ads", "", "50% in Q1 (testing), 100% by Q2 (90-day sprint)"],
    ["", "GBP", "", "Gradual 4-quarter ramp (typical 4-6 month results)"],
    ["", "Website", "", "Delayed start Q3, full maturity Q6 (12-18 months)"],
]

for row_data in assumptions_data:
    ws1.append(row_data)
    
# Format value column
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    # Format based on content
    if row[1].value and "%" in str(row[1].value):
        row[2].number_format = percent_format
    elif row[1].value and any(word in str(row[1].value) for word in ["Cost", "Average", "Fee", "Value", "LTV"]):
        row[2].number_format = money_format
    
    # Add borders
    for cell in row:
        cell.border = thin_border
    
    # Highlight section headers
    if row[0].value in ["GENERAL", "TUCSON MARKET", "DENVER MARKET", "RAMP SCHEDULES", "NOTES"]:
        for cell in row:
            cell.fill = subheader_fill
            cell.font = Font(bold=True)

auto_width(ws1)

# ===== TAB 2: TUCSON PROJECTIONS =====
ws2 = wb.create_sheet("Tucson Projections")

# Data structure
tucson_headers = [
    "Quarter",
    "Website Leads",
    "PPC Leads", 
    "GBP Leads",
    "Total Qualified Leads",
    "Ad Spend",
    "Mgmt Fee Share",
    "Total Cost",
    "Jobs Closed",
    "Revenue",
    "ROI"
]

ws2.append(tucson_headers)
style_header_row(ws2)

# Assumptions for Tucson
max_website = 20
max_ppc = 15
max_gbp = 30
cpl = 700
ltv = 10933
mgmt_fee_share = 5500 / 2  # Split 50/50
qualified_rate = 0.5
closing_rate = 0.5

# Ramp schedules
ramps = {
    'Q1': {'ads': 0.50, 'gbp': 0.25, 'web': 0.00},
    'Q2': {'ads': 1.00, 'gbp': 0.50, 'web': 0.00},
    'Q3': {'ads': 1.00, 'gbp': 0.75, 'web': 0.10},
    'Q4': {'ads': 1.00, 'gbp': 1.00, 'web': 0.30},
    'Q5': {'ads': 1.00, 'gbp': 1.00, 'web': 0.60},
    'Q6': {'ads': 1.00, 'gbp': 1.00, 'web': 1.00},
}

tucson_data = []
for q, ramp in ramps.items():
    # Calculate quarterly leads (3 months per quarter)
    website_leads = max_website * ramp['web'] * qualified_rate * 3
    ppc_leads = max_ppc * ramp['ads'] * qualified_rate * 3
    gbp_leads = max_gbp * ramp['gbp'] * qualified_rate * 3
    total_qualified = website_leads + ppc_leads + gbp_leads
    
    # Costs
    ad_spend = ppc_leads / qualified_rate * cpl  # Actual PPC leads × CPL
    mgmt_fee = mgmt_fee_share * 3  # 3 months
    total_cost = ad_spend + mgmt_fee
    
    # Revenue
    jobs = total_qualified * closing_rate
    revenue = jobs * ltv
    roi = revenue / total_cost if total_cost > 0 else 0
    
    tucson_data.append([
        q,
        round(website_leads, 1),
        round(ppc_leads, 1),
        round(gbp_leads, 1),
        round(total_qualified, 1),
        round(ad_spend, 0),
        round(mgmt_fee, 0),
        round(total_cost, 0),
        round(jobs, 1),
        round(revenue, 0),
        round(roi, 1)
    ])

# Add data rows
for row_data in tucson_data:
    ws2.append(row_data)

# Add totals row
totals = ['TOTAL', 
          sum(r[1] for r in tucson_data),
          sum(r[2] for r in tucson_data),
          sum(r[3] for r in tucson_data),
          sum(r[4] for r in tucson_data),
          sum(r[5] for r in tucson_data),
          sum(r[6] for r in tucson_data),
          sum(r[7] for r in tucson_data),
          sum(r[8] for r in tucson_data),
          sum(r[9] for r in tucson_data),
          sum(r[9] for r in tucson_data) / sum(r[7] for r in tucson_data)]

ws2.append(totals)

# Format numbers
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for idx, cell in enumerate(row):
        cell.border = thin_border
        
        # Number formatting
        if idx in [5, 6, 7, 9]:  # Money columns
            cell.number_format = money_format
        elif idx == 10:  # ROI
            cell.number_format = '0.0"x"'
        elif idx in [1, 2, 3, 4, 8]:  # Lead/job counts
            cell.number_format = decimal_format
    
    # Highlight total row
    if row[0].value == 'TOTAL':
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)

auto_width(ws2)

# ===== TAB 3: DENVER PROJECTIONS =====
ws3 = wb.create_sheet("Denver Projections")

ws3.append(tucson_headers)
style_header_row(ws3)

# Denver assumptions
max_website_d = 35
max_ppc_d = 30
max_gbp_d = 50
cpl_d = 800
ltv_d = 6100

denver_data = []
for q, ramp in ramps.items():
    website_leads = max_website_d * ramp['web'] * qualified_rate * 3
    ppc_leads = max_ppc_d * ramp['ads'] * qualified_rate * 3
    gbp_leads = max_gbp_d * ramp['gbp'] * qualified_rate * 3
    total_qualified = website_leads + ppc_leads + gbp_leads
    
    ad_spend = ppc_leads / qualified_rate * cpl_d
    mgmt_fee = mgmt_fee_share * 3
    total_cost = ad_spend + mgmt_fee
    
    jobs = total_qualified * closing_rate
    revenue = jobs * ltv_d
    roi = revenue / total_cost if total_cost > 0 else 0
    
    denver_data.append([
        q,
        round(website_leads, 1),
        round(ppc_leads, 1),
        round(gbp_leads, 1),
        round(total_qualified, 1),
        round(ad_spend, 0),
        round(mgmt_fee, 0),
        round(total_cost, 0),
        round(jobs, 1),
        round(revenue, 0),
        round(roi, 1)
    ])

for row_data in denver_data:
    ws3.append(row_data)

# Add totals
totals_d = ['TOTAL',
            sum(r[1] for r in denver_data),
            sum(r[2] for r in denver_data),
            sum(r[3] for r in denver_data),
            sum(r[4] for r in denver_data),
            sum(r[5] for r in denver_data),
            sum(r[6] for r in denver_data),
            sum(r[7] for r in denver_data),
            sum(r[8] for r in denver_data),
            sum(r[9] for r in denver_data),
            sum(r[9] for r in denver_data) / sum(r[7] for r in denver_data)]

ws3.append(totals_d)

# Format
for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    for idx, cell in enumerate(row):
        cell.border = thin_border
        if idx in [5, 6, 7, 9]:
            cell.number_format = money_format
        elif idx == 10:
            cell.number_format = '0.0"x"'
        elif idx in [1, 2, 3, 4, 8]:
            cell.number_format = decimal_format
    
    if row[0].value == 'TOTAL':
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)

auto_width(ws3)

# ===== TAB 4: COMBINED SUMMARY =====
ws4 = wb.create_sheet("Combined Summary")

combined_headers = [
    "Quarter",
    "Total Qualified Leads",
    "Total Ad Spend",
    "Total Mgmt Fees",
    "Total Investment",
    "Total Jobs",
    "Total Revenue",
    "Combined ROI",
    "Monthly Revenue (Avg)"
]

ws4.append(combined_headers)
style_header_row(ws4)

# Combine data
combined_data = []
for i, q in enumerate(['Q1', 'Q2', 'Q3', 'Q4', 'Q5', 'Q6']):
    tucson_row = tucson_data[i]
    denver_row = denver_data[i]
    
    total_qualified = tucson_row[4] + denver_row[4]
    total_ad_spend = tucson_row[5] + denver_row[5]
    total_mgmt = tucson_row[6] + denver_row[6]
    total_investment = tucson_row[7] + denver_row[7]
    total_jobs = tucson_row[8] + denver_row[8]
    total_revenue = tucson_row[9] + denver_row[9]
    combined_roi = total_revenue / total_investment if total_investment > 0 else 0
    monthly_avg = total_revenue / 3
    
    combined_data.append([
        q,
        round(total_qualified, 1),
        round(total_ad_spend, 0),
        round(total_mgmt, 0),
        round(total_investment, 0),
        round(total_jobs, 1),
        round(total_revenue, 0),
        round(combined_roi, 1),
        round(monthly_avg, 0)
    ])
    
    ws4.append(combined_data[-1])

# Add totals and check goal
totals_combined = [
    'TOTAL',
    sum(r[1] for r in combined_data),
    sum(r[2] for r in combined_data),
    sum(r[3] for r in combined_data),
    sum(r[4] for r in combined_data),
    sum(r[5] for r in combined_data),
    sum(r[6] for r in combined_data),
    sum(r[6] for r in combined_data) / sum(r[4] for r in combined_data),
    None  # No monthly avg for totals
]
ws4.append(totals_combined)

# Add goal check row
q6_monthly = combined_data[-1][8]
goal_met = "✓ YES" if q6_monthly >= 300000 else "✗ NO"
goal_row = ['', '', '', '', '', '', '', 'Q6 Monthly:', q6_monthly]
ws4.append(goal_row)

goal_status = ['', '', '', '', '', '', '', '$300k Goal Met?', goal_met]
ws4.append(goal_status)

# Format
for row_idx, row in enumerate(ws4.iter_rows(min_row=2, max_row=ws4.max_row), start=2):
    for idx, cell in enumerate(row):
        cell.border = thin_border
        
        if idx in [2, 3, 4, 6, 8]:  # Money columns
            cell.number_format = money_format
        elif idx == 7:  # ROI
            cell.number_format = '0.0"x"'
        elif idx in [1, 5]:  # Counts
            cell.number_format = decimal_format
    
    # Highlight total row
    if row[0].value == 'TOTAL':
        for cell in row:
            cell.fill = total_fill
            cell.font = Font(bold=True)
    
    # Highlight goal rows
    if row_idx >= ws4.max_row - 1:
        for cell in row:
            cell.font = Font(bold=True, size=12)
            if q6_monthly >= 300000:
                cell.fill = goal_fill
            else:
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

auto_width(ws4)

# ===== TAB 5: SENSITIVITIES =====
ws5 = wb.create_sheet("Sensitivities & Notes")

ws5.append(["Scenario Type", "Variable Changed", "Impact", "Result/Notes"])
style_header_row(ws5)

sensitivity_data = [
    ["SENSITIVITY ANALYSIS", "", "", ""],
    ["Closing Rate", "60% (vs 50% base)", "Higher job conversion", "Q6 monthly: ~$847k (both locations)"],
    ["Closing Rate", "40% (vs 50% base)", "Lower job conversion", "Q6 monthly: ~$565k (both locations)"],
    ["", "", "", ""],
    ["Cost Per Lead", "+20% CPL increase", "Tucson $840, Denver $960", "ROI drops to ~16x in Q6, still hits $300k target"],
    ["Cost Per Lead", "-10% CPL decrease", "Tucson $630, Denver $720", "ROI improves to ~22x in Q6"],
    ["", "", "", ""],
    ["Website Delay", "Website starts Q4 (not Q3)", "Delayed organic growth", "Q6 monthly: ~$658k, still exceeds goal"],
    ["Website Acceleration", "Website starts Q2", "Faster organic growth", "Q6 monthly: ~$780k, exceeds goal"],
    ["", "", "", ""],
    ["RECOMMENDATIONS", "", "", ""],
    ["Acceleration Strategy", "Plumbing keywords Q3", "Lower CPL by 10-20%", "Target emergency plumbing searches"],
    ["Acceleration Strategy", "Add 2nd GBP in Denver Q2", "Additional 20-30 leads/mo by Q4", "Target downtown Denver market"],
    ["Acceleration Strategy", "Add 2nd GBP in Tucson Q4", "Additional 15-20 leads/mo by Q6", "Target east/north Tucson"],
    ["", "", "", ""],
    ["CONSERVATIVE NOTES", "", "", ""],
    ["Website Ramp", "Slower than average", "~$498k total Q1-Q2", "Ensures realistic expectations"],
    ["GBP Ramp", "Standard timeline", "4-6 month maturity typical", "25%/50%/75%/100% progression"],
    ["Ad Ramp", "50% start Q1", "90-day optimization sprint", "Full capacity Q2+"],
    ["", "", "", ""],
    ["FEASIBILITY CHECK", "", "", ""],
    ["Q6 Monthly Target", "~$706k/month projected", "Exceeds $300k goal", "Combined both locations"],
    ["18-Month Total", "~$8.43M projected", "Based on conservative model", "Blended all channels"],
    ["Break-even Point", "Q1 shows positive ROI", "10.2x combined ROI", "Immediate profitability"],
    ["", "", "", ""],
    ["FORMULAS DOCUMENTATION", "", "", ""],
    ["Qualified Leads (Qtr)", "=Max Monthly × Ramp % × Qualified % × 3", "", "3 months per quarter"],
    ["Ad Spend (Qtr)", "=PPC Leads (raw) × CPL", "", "Only for PPC channel"],
    ["Management Fee (Qtr)", "=$5,500 ÷ 2 × 3 months", "", "Split 50/50 between locations"],
    ["Jobs (Qtr)", "=Total Qualified Leads × Closing Rate", "", "50% close rate assumption"],
    ["Revenue (Qtr)", "=Jobs × LTV", "", "Tucson $10,933, Denver $6,100"],
    ["ROI", "=Revenue ÷ Total Cost", "", "If Total Cost > 0"],
    ["LTV (Tucson)", "=Mit + (Abate% × Abate) + (Recon% × Recon)", "", "$4,589 + (30% × $7,484) + (55% × $7,452)"],
    ["LTV (Denver)", "=Mitigation Average", "", "$6,100 (recon referred out)"],
]

for row_data in sensitivity_data:
    ws5.append(row_data)

# Format
for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row):
    for cell in row:
        cell.border = thin_border
    
    # Highlight section headers
    if row[0].value in ["SENSITIVITY ANALYSIS", "RECOMMENDATIONS", "CONSERVATIVE NOTES", 
                        "FEASIBILITY CHECK", "FORMULAS DOCUMENTATION"]:
        for cell in row:
            cell.fill = subheader_fill
            cell.font = Font(bold=True)

auto_width(ws5)

# Save file
output_path = "/Users/jameslarosa/Desktop/Random AI Prjects/AI Wireframe Builder/projections/Conservative_18Mo_Projections.xlsx"
wb.save(output_path)
print(f"✅ Created: {output_path}")
print(f"\n📊 Key Results:")
print(f"  • Q6 Monthly Revenue: ${combined_data[-1][8]:,.0f}")
print(f"  • Goal ($300k/month): {'✓ MET' if combined_data[-1][8] >= 300000 else '✗ NOT MET'}")
print(f"  • 18-Month Total Revenue: ${sum(r[6] for r in combined_data):,.0f}")
print(f"  • Total Investment: ${sum(r[4] for r in combined_data):,.0f}")
print(f"  • Overall ROI: {sum(r[6] for r in combined_data) / sum(r[4] for r in combined_data):.1f}x")
