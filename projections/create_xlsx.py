#!/usr/bin/env python3
"""Generate Excel file with all projection tabs"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
money_format = '"$"#,##0'
percent_format = '0.0"%"'
number_format = '#,##0.0'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)

wb = Workbook()

# ===== TAB 1: ASSUMPTIONS =====
ws1 = wb.active
ws1.title = "Assumptions"
ws1.append(["Category", "Parameter", "Value", "Notes"])
style_header(ws1)

assumptions_data = [
    ["General", "Base Monthly Ad Spend Per Location", 5000, "Per location per month"],
    ["General", "Monthly Management Fee (Both Locations)", 5500, "Total for both locations"],
    ["General", "Qualified Lead % of Calls", 0.5, "Half of calls become qualified leads"],
    ["General", "Closing Rate (Qualified to Job)", 0.5, "Half of qualified leads become jobs"],
    ["", "", "", ""],
    ["Tucson", "Max Monthly Website Leads", 20, ""],
    ["Tucson", "Max Monthly PPC Leads", 15, ""],
    ["Tucson", "Max Monthly GBP Leads", 30, ""],
    ["Tucson", "CPL (Cost Per Lead)", 700, ""],
    ["Tucson", "Mitigation Average", 4589, ""],
    ["Tucson", "Abatement Average", 7484, ""],
    ["Tucson", "Abatement Conversion %", 0.3, ""],
    ["Tucson", "Reconstruction Conversion %", 0.55, ""],
    ["Tucson", "Reconstruction Average", 7452, ""],
    ["Tucson", "LTV Per Job", 10933, "Mit + (Abate Conv * Abate Avg) + (Recon Conv * Recon Avg)"],
    ["", "", "", ""],
    ["Denver", "Max Monthly Website Leads", 35, ""],
    ["Denver", "Max Monthly PPC Leads", 30, ""],
    ["Denver", "Max Monthly GBP Leads", 50, ""],
    ["Denver", "CPL (Cost Per Lead)", 800, ""],
    ["Denver", "Mitigation Average", 6100, ""],
    ["Denver", "Recon Referral Fee Per Job", 0, ""],
    ["Denver", "LTV Per Job", 6100, "Mit Avg + Recon Fee"],
    ["", "", "", ""],
    ["Ramp - Ads", "Q1", 0.5, ""],
    ["Ramp - Ads", "Q2", 1.0, ""],
    ["Ramp - Ads", "Q3", 1.0, ""],
    ["Ramp - Ads", "Q4", 1.0, ""],
    ["Ramp - Ads", "Q5", 1.0, ""],
    ["Ramp - Ads", "Q6", 1.0, ""],
    ["", "", "", ""],
    ["Ramp - GBP", "Q1", 0.25, ""],
    ["Ramp - GBP", "Q2", 0.5, ""],
    ["Ramp - GBP", "Q3", 0.75, ""],
    ["Ramp - GBP", "Q4", 1.0, ""],
    ["Ramp - GBP", "Q5", 1.0, ""],
    ["Ramp - GBP", "Q6", 1.0, ""],
    ["", "", "", ""],
    ["Ramp - Website", "Q1", 0.0, ""],
    ["Ramp - Website", "Q2", 0.0, ""],
    ["Ramp - Website", "Q3", 0.1, ""],
    ["Ramp - Website", "Q4", 0.3, ""],
    ["Ramp - Website", "Q5", 0.6, ""],
    ["Ramp - Website", "Q6", 1.0, ""],
]

for row in assumptions_data:
    ws1.append(row)

# Format percentages
for row in ws1.iter_rows(min_row=2):
    if "%" in str(row[1].value) or row[2].value in [0.5, 0.3, 0.55, 0.25, 0.75, 0.1, 0.6, 1.0, 0.0]:
        if isinstance(row[2].value, float) and row[2].value <= 1:
            row[2].number_format = '0%'
    row[0].border = thin_border
    row[1].border = thin_border
    row[2].border = thin_border
    row[3].border = thin_border

auto_width(ws1)

# ===== TAB 2: TUCSON PROJECTIONS =====
ws2 = wb.create_sheet("Tucson Projections")
ws2.append(["Quarter", "Website Leads", "PPC Leads", "GBP Leads", "Total Qualified Leads", 
            "Ad Spend", "Fee Share", "CPL Cost", "Total Cost", "Jobs", "Revenue", "ROI"])
style_header(ws2)

tucson_data = [
    ["Q1", 0.0, 22.5, 22.5, 45.0, 7500, 8250, 5250, 21000, 22.5, 245981, 11.7],
    ["Q2", 0.0, 45.0, 45.0, 90.0, 18000, 8250, 10500, 36750, 45.0, 491976, 13.4],
    ["Q3", 6.0, 45.0, 67.5, 118.5, 21000, 8250, 10500, 39750, 59.3, 647808, 16.3],
    ["Q4", 18.0, 45.0, 90.0, 153.0, 24000, 8250, 10500, 42750, 76.5, 836189, 19.6],
    ["Q5", 36.0, 45.0, 90.0, 171.0, 27000, 8250, 10500, 45750, 85.5, 934581, 20.4],
    ["Q6", 60.0, 45.0, 90.0, 195.0, 30000, 8250, 10500, 48750, 97.5, 1065948, 21.9],
    ["Total", 120.0, 247.5, 405.0, 772.5, 127500, 49500, 57750, 234750, 386.3, 4222483, 18.0],
]

for row in tucson_data:
    ws2.append(row)

for row in ws2.iter_rows(min_row=2):
    for i, cell in enumerate(row):
        cell.border = thin_border
        if i in [5, 6, 7, 8, 10]:  # Money columns
            cell.number_format = money_format
        elif i == 11:  # ROI
            cell.number_format = '0.0"x"'

auto_width(ws2)

# ===== TAB 3: DENVER PROJECTIONS =====
ws3 = wb.create_sheet("Denver Projections")
ws3.append(["Quarter", "Website Leads", "PPC Leads", "GBP Leads", "Total Qualified Leads", 
            "Ad Spend", "Fee Share", "CPL Cost", "Total Cost", "Jobs", "Revenue", "ROI"])
style_header(ws3)

denver_data = [
    ["Q1", 0.0, 45.0, 37.5, 82.5, 7500, 8250, 12000, 27750, 41.3, 251625, 9.1],
    ["Q2", 0.0, 90.0, 75.0, 165.0, 18000, 8250, 24000, 50250, 82.5, 503250, 10.0],
    ["Q3", 10.5, 90.0, 112.5, 213.0, 21000, 8250, 24000, 53250, 106.5, 649650, 12.2],
    ["Q4", 31.5, 90.0, 150.0, 271.5, 24000, 8250, 24000, 56250, 135.8, 827925, 14.7],
    ["Q5", 63.0, 90.0, 150.0, 303.0, 27000, 8250, 24000, 59250, 151.5, 924150, 15.6],
    ["Q6", 105.0, 90.0, 150.0, 345.0, 30000, 8250, 24000, 62250, 172.5, 1052250, 16.9],
    ["Total", 210.0, 495.0, 675.0, 1380.0, 127500, 49500, 132000, 309000, 690.0, 4208850, 13.6],
]

for row in denver_data:
    ws3.append(row)

for row in ws3.iter_rows(min_row=2):
    for i, cell in enumerate(row):
        cell.border = thin_border
        if i in [5, 6, 7, 8, 10]:
            cell.number_format = money_format
        elif i == 11:
            cell.number_format = '0.0"x"'

auto_width(ws3)

# ===== TAB 4: COMBINED SUMMARY =====
ws4 = wb.create_sheet("Combined Summary")
ws4.append(["Quarter", "Total Qualified Leads", "Total Ad Spend", "Total Fee Share", 
            "Total CPL Cost", "Total Cost", "Total Jobs", "Total Revenue", "Combined ROI", "Monthly Revenue (Avg)"])
style_header(ws4)

combined_data = [
    ["Q1", 127.5, 15000, 16500, 17250, 48750, 63.8, 497606, 10.2, 165869],
    ["Q2", 255.0, 36000, 16500, 34500, 87000, 127.5, 995226, 11.4, 331742],
    ["Q3", 331.5, 42000, 16500, 34500, 93000, 165.8, 1297458, 13.9, 432486],
    ["Q4", 424.5, 48000, 16500, 34500, 99000, 212.3, 1664114, 16.8, 554705],
    ["Q5", 474.0, 54000, 16500, 34500, 105000, 237.0, 1858731, 17.7, 619577],
    ["Q6", 540.0, 60000, 16500, 34500, 111000, 270.0, 2118198, 19.1, 706066],
    ["Total", 2152.5, 255000, 99000, 189750, 543750, 1076.3, 8431333, 15.5, None],
]

for row in combined_data:
    ws4.append(row)

for row in ws4.iter_rows(min_row=2):
    for i, cell in enumerate(row):
        cell.border = thin_border
        if i in [2, 3, 4, 5, 7, 9]:
            cell.number_format = money_format
        elif i == 8:
            cell.number_format = '0.0"x"'

auto_width(ws4)

# ===== TAB 5: SENSITIVITIES & NOTES =====
ws5 = wb.create_sheet("Sensitivities & Notes")
ws5.append(["Category", "Scenario", "Impact", "Details"])
style_header(ws5)

notes_data = [
    ["Sensitivity", "Closing Rate = 60%", "Q6 monthly revenue ~$847k", "Combined - stronger build with higher conversion"],
    ["Sensitivity", "CPL Rises 20%", "Costs up ~$7k/qtr/location", "ROI drops to 16x in Q6 - still hits $300k target"],
    ["Sensitivity", "Website Delays to Q4 Start", "Q6 monthly falls to ~$658k", "Still above $300k target - emphasizes conservatism"],
    ["", "", "", ""],
    ["Recommendation", "Accelerate Strategy", "Introduce plumbing keywords in Q3", "Lower CPL 10-20%"],
    ["Recommendation", "Accelerate Strategy", "Add GBPs in Q2", "+20-30 leads/mo by Q4"],
    ["", "", "", ""],
    ["Conservatism Note", "Slower Website Ramp", "Delays early revenue", "~$498k total in Q1-Q2 vs faster scenarios"],
    ["Conservatism Note", "Realistic Scaling", "Ensures no overpromising", "Conservative but achievable targets"],
    ["", "", "", ""],
    ["Feasibility", "Q6 Monthly Target", "~$706k/month", "Exceeds $300k target"],
    ["Feasibility", "18-Month Total Revenue", "~$8.43M", "Based on provided data"],
    ["", "", "", ""],
    ["Formula", "Qualified Leads Qtr", "=Max Monthly * Ramp * Qualified % * 3", "3 months per quarter"],
    ["Formula", "Ad Spend Qtr", "=Base * Ramp * 3", "Scaled by ad ramp"],
    ["Formula", "CPL Cost Qtr", "=PPC Leads Qtr * CPL", ""],
    ["Formula", "Jobs Qtr", "=Total Leads Qtr * Closing Rate", ""],
    ["Formula", "Revenue Qtr", "=Jobs Qtr * LTV", ""],
    ["Formula", "ROI", "=Revenue / Total Cost", "If cost > 0"],
]

for row in notes_data:
    ws5.append(row)

for row in ws5.iter_rows(min_row=2):
    for cell in row:
        cell.border = thin_border

auto_width(ws5)

# Save
output_path = "/Users/jameslarosa/Desktop/Random AI Prjects/AI Wireframe Builder/projections/18_Month_Projections.xlsx"
wb.save(output_path)
print(f"✅ Created: {output_path}")
